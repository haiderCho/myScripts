#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import csv
import openpyxl
import sys
import argparse
from pathlib import Path
from openpyxl import Workbook


def csv_to_excel(csv_file, excel_file, sheet_name='Sheet1', separator=',', start_row=1, start_col=1):
    """
    Convert CSV file to Excel
    
    Args:
        csv_file: Path to input CSV file
        excel_file: Path to output Excel file
        sheet_name: Name of the Excel sheet
        separator: CSV delimiter
        start_row: Starting row in Excel (1-indexed)
        start_col: Starting column in Excel (1-indexed)
    """
    csv_path = Path(csv_file)
    excel_path = Path(excel_file)
    
    # Validate CSV exists
    if not csv_path.exists():
        raise FileNotFoundError(f"CSV file not found: {csv_file}")
    
    # Load or create workbook
    if excel_path.exists():
        print(f"Loading existing Excel file: {excel_file}")
        wb = openpyxl.load_workbook(excel_file)
    else:
        print(f"Creating new Excel file: {excel_file}")
        wb = Workbook()
        # Remove default sheet if creating new workbook
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
    
    # Create or select sheet
    if sheet_name in wb.sheetnames:
        print(f"Writing to existing sheet: {sheet_name}")
        sheet = wb[sheet_name]
    else:
        print(f"Creating new sheet: {sheet_name}")
        sheet = wb.create_sheet(title=sheet_name)
    
    # Read CSV and write to Excel
    with open(csv_path, "r", encoding="utf-8", newline="") as f:
        reader = csv.reader(f, delimiter=separator)
        
        row_count = 0
        for row_index, row_data in enumerate(reader, start=start_row):
            for col_index, value in enumerate(row_data, start=start_col):
                # Try to convert numbers
                try:
                    if '.' in value:
                        cell_value = float(value)
                    else:
                        cell_value = int(value)
                except (ValueError, AttributeError):
                    cell_value = value
                
                sheet.cell(row=row_index, column=col_index, value=cell_value)
            row_count += 1
    
    wb.save(excel_file)
    print(f"✓ Successfully wrote {row_count} rows to '{sheet_name}' in {excel_file}")


def batch_csv_to_excel(csv_files, output_dir=None, separator=','):
    """Convert multiple CSV files to Excel"""
    for csv_file in csv_files:
        csv_path = Path(csv_file)
        
        if not csv_path.exists():
            print(f"⚠️  Skipping {csv_file}: File not found")
            continue
        
        # Determine output path
        if output_dir:
            out_dir = Path(output_dir)
            out_dir.mkdir(parents=True, exist_ok=True)
            excel_file = out_dir / f"{csv_path.stem}.xlsx"
        else:
            excel_file = csv_path.with_suffix('.xlsx')
        
        try:
            print(f"\nProcessing: {csv_path.name}")
            csv_to_excel(csv_file, excel_file, sheet_name=csv_path.stem, separator=separator)
        except Exception as e:
            print(f"✗ Error processing {csv_file}: {e}")


def main():
    parser = argparse.ArgumentParser(
        description='Convert CSV files to Excel format with advanced options',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
Examples:
  # Interactive mode
  %(prog)s
  
  # Single file conversion
  %(prog)s data.csv -o output.xlsx
  
  # With custom separator and sheet name
  %(prog)s data.csv -o output.xlsx --sep ";" --sheet "MyData"
  
  # Batch conversion
  %(prog)s file1.csv file2.csv file3.csv --batch --output-dir excel_files
  
  # Start writing at specific cell
  %(prog)s data.csv -o output.xlsx --start-row 5 --start-col 3
        '''
    )
    
    parser.add_argument('input', nargs='*', help='CSV file(s) to convert')
    parser.add_argument('-o', '--output', help='Output Excel file path')
    parser.add_argument('--sep', '--separator', default=',', help='CSV separator (default: comma)')
    parser.add_argument('--sheet', default='Sheet1', help='Excel sheet name (default: Sheet1)')
    parser.add_argument('--start-row', type=int, default=1, help='Starting row in Excel (default: 1)')
    parser.add_argument('--start-col', type=int, default=1, help='Starting column in Excel (default: 1)')
    parser.add_argument('--batch', action='store_true', help='Batch mode: convert multiple CSV files')
    parser.add_argument('--output-dir', help='Output directory for batch mode')
    
    args = parser.parse_args()
    
    try:
        # Interactive mode
        if not args.input:
            print("CSV to Excel Converter")
            print("=" * 50)
            print("Input and output files can be absolute or relative paths.\n")
            
            csv_file = input("Name of the CSV file for input: ").strip()
            sep = input("Separator used in the CSV file (default ,): ").strip() or ','
            excel_file = input("Name of the Excel file for output: ").strip()
            sheet_name = input("Name of the Excel sheet (default Sheet1): ").strip() or 'Sheet1'
            
            csv_to_excel(csv_file, excel_file, sheet_name, sep)
        
        # Batch mode
        elif args.batch:
            batch_csv_to_excel(args.input, args.output_dir, args.sep)
        
        # Single file mode
        else:
            if len(args.input) > 1:
                print("Error: Specify only one input file, or use --batch for multiple files")
                sys.exit(1)
            
            csv_file = args.input[0]
            
            # Determine output file
            if args.output:
                excel_file = args.output
            else:
                excel_file = Path(csv_file).with_suffix('.xlsx')
            
            csv_to_excel(csv_file, excel_file, args.sheet, args.sep, args.start_row, args.start_col)
    
    except FileNotFoundError as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
